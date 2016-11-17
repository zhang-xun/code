#! /usr/local/bin/python3
# _*_ coding:utf-8 _*_

from openpyxl import Workbook
from openpyxl import load_workbook
wb = load_workbook(filename= 'tobeanalysis.xlsm')
sheet_range = wb['Sheet1']
print(sheet_range['C68'].value)
print(sheet_range.cell(row=68,column=3).value)


def onecell_iter():
    for i in range(2,138):
        for j in range(1,7):
            print(sheet_range.cell(row=i,column=j).value)


def manycell_iter():
    for row in sheet_range.iter_rows(min_row=2,max_col=6,max_row=138):
        for cell in row:
            print(cell.value)

if __name__ == "__main__":
    #onecell_iter()
    manycell_iter()
